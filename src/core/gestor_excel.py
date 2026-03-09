import os
import json
from openpyxl import Workbook, load_workbook


def obtener_ruta_excel():
    ruta_escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
    return os.path.join(ruta_escritorio, "registro_facturas.xlsx")


def factura_existe_en_excel(proveedor, numero_factura):
    """Verifica si la factura ya está registrada para evitar duplicados contables."""
    if not proveedor or not numero_factura:
        return False  # Si faltan datos clave, no podemos asegurar que sea duplicada

    ruta_excel = obtener_ruta_excel()
    if not os.path.isfile(ruta_excel):
        return False

    libro = load_workbook(ruta_excel, read_only=True)
    hoja = libro.active

    # Obtenemos los nombres de las columnas
    cabeceras = [celda.value for celda in hoja[1]]

    try:
        idx_prov = cabeceras.index("proveedor_emisor")
        idx_num = cabeceras.index("numero_factura")
    except ValueError:
        libro.close()
        return False

    # Comprobamos fila a fila
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        # Normalizamos a minúsculas y quitamos espacios extra por seguridad
        prov_excel = str(fila[idx_prov]).strip().lower() if fila[idx_prov] else ""
        num_excel = str(fila[idx_num]).strip().lower() if fila[idx_num] else ""

        prov_actual = str(proveedor).strip().lower()
        num_actual = str(numero_factura).strip().lower()

        if prov_excel == prov_actual and num_excel == num_actual:
            libro.close()
            return True  # Encontramos el duplicado

    libro.close()
    return False


def guardar_en_excel(datos_factura):
    ruta_excel = obtener_ruta_excel()

    if 'lineas_iva' in datos_factura and isinstance(datos_factura['lineas_iva'], list):
        datos_factura['lineas_iva'] = json.dumps(datos_factura['lineas_iva'])

    cabeceras = [
        "nombre_documento", "proveedor_emisor", "cliente_receptor", "numero_factura",
        "fecha", "total_base_imponible", "total_cuota_iva", "total", "lineas_iva"
    ]

    if not os.path.isfile(ruta_excel):
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Facturas Extraidas"
        hoja.append(cabeceras)
    else:
        libro = load_workbook(ruta_excel)
        hoja = libro.active

    fila_datos = []
    for columna in cabeceras:
        valor = datos_factura.get(columna, "")
        if valor is None:
            valor = ""
        fila_datos.append(valor)

    hoja.append(fila_datos)
    libro.save(ruta_excel)

    return ruta_excel